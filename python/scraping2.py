import requests
from datetime import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from time import sleep
navegador = webdriver.Chrome(executable_path= r"D:\Users\t_flaviomendes.JFCE\Downloads\python\chromedriver")
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook

def baterponto ():
    ########### LOGANDO NO PONTO ##################################################

    navegador.get('https://app2.pontomais.com.br/registrar-ponto')

    navegador.implicitly_wait(30)
    #sleep(7)

    login = navegador.find_elements_by_tag_name('input')[0]
    login.send_keys('04508536328')

    #sleep(1)

    passwd = navegador.find_elements_by_tag_name('input')[1]
    passwd.send_keys('flavio144F!@')

    #sleep(1)

    passwd.send_keys(Keys.ENTER)

    ########## FIM LOGANDO NO PONTO ####################################################

    #sleep(10)

    ######### VERIFICANDO ÚLTIMO REGISTRO ##############################################

    ultimo_registro = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[1]/last-time-card/div/div/p').text
    print(ultimo_registro)
    ######### FIM DO VERIFICANDO ÚLTIMO REGISTRO #######################################

    #sleep(2)

    ######## BATENDO O PONTO ###########################################################    
    
    localizacao = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[2]/address-time-card-register/div[2]/p[3]/small')

    #sleep(2)

    localizacao.click()

    #sleep(3)

    bater = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[2]/div/pm-button/button')
    bater.click()

    ######## FIM DO BATENDO PONTO #########################################################

    #sleep(2)

    ######## VERIFICANDO PONTO #############################################################

    navegador.refresh()

    #sleep(3)

    agora = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[1]/last-time-card/div/div/p').text
    print(agora)

    if agora == ultimo_registro:
        print("Não bateu o Ponto")
        baterPontoNovamente()
    else:

        print("Ponto Registrado com sucesso.")   

    navegador.quit()     

    ####### FIM DO VERIFICANDO PONTO ########################################################

    # data_e_hora_atuais = datetime.now()

    # agora = data_e_hora_atuais.strftime('%d/%m às %H:%M')
    
    # if agora == ponto:
    #     print("bateu")
        
    # else:
    #     baterponto()


def baterPontoNovamente():

    navegador.get('https://app2.pontomais.com.br/registrar-ponto')
    navegador.implicitly_wait(30)

    #sleep(7)

    ######### VERIFICANDO ÚLTIMO REGISTRO ##############################################

    ultimo_registro = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[1]/last-time-card/div/div/p').text
    print(ultimo_registro)

    ######### FIM DO VERIFICANDO ÚLTIMO REGISTRO #######################################

    #sleep(2)

    ######## BATENDO O PONTO ###########################################################    
    
    localizacao = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[2]/address-time-card-register/div[2]/p[3]/small')

    #sleep(2)

    localizacao.click()

    #sleep(3)

    bater = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[2]/div/pm-button/button')
    bater.click()

    ######## FIM DO BATENDO PONTO #########################################################

    #sleep(2)

    ######## VERIFICANDO PONTO #############################################################

    navegador.refresh()

    #sleep(3)
    
    agora = navegador.find_element_by_xpath('/html/body/app-root/app-side-nav-outer-toolbar/dx-drawer/div/div[2]/dx-scroll-view/div[1]/div/div[1]/div[2]/div[1]/app-container/time-card-register/div/div[2]/div[2]/pm-time-card-register/pm-card/div/div[2]/div[1]/div[1]/last-time-card/div/div/p').text
    print(agora)

    if agora == ultimo_registro:
        print("Não bateu ponto 2")
        baterPontoNovamente()
    else:

        print("Ponto Registrado com sucesso.") 

    navegador.quit()
    # wb = load_workbook("ponto.xlsx")
    # ws = wb.active

    # ws['A'] = agora

    # wb.save()

baterponto()




    







